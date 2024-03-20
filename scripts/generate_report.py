#!/opt/netbox/venv/bin/python
from os import environ, getcwd, makedirs
from os import path as os_path
from django import setup
from sys import exc_info, path
path.append('/opt/netbox/netbox')
environ.setdefault('DJANGO_SETTINGS_MODULE', 'netbox.settings')
setup()
# CLI Run with:
# source /opt/netbox/venv/bin/activate
# python manage.py runscript --loglevel debug --commit --data '{"limit": 5000, "offset": 1, "base_url": "10.0.1.1:8080", "username": "test_user", "password": "mypassword123", "devices": "", "with_logs": true, "nso_timeout": 500, "with_nso": false, "nso_retry": 1, "with_multithreading": true}' generate_report.GenerateReport

from extras.scripts import Script, StringVar, TextVar, IntegerVar, BooleanVar
from django.forms import PasswordInput
from datetime import datetime
from utilities.exceptions import AbortScript
from dcim.models import Interface
from django.core.exceptions import ValidationError
from concurrent.futures import ThreadPoolExecutor
from threading import BoundedSemaphore
from traceback import format_exc
from requests.exceptions import Timeout as TimeoutException
from requests.exceptions import ConnectionError




"""
requirements.txt:
pip install -i http://jfrog.local/artifactory/api/pypi/pypi/simple --trusted-host jfrog.local openpyxl==3.1.2
"""

# # Create a semaphore for the NCS requests
# ncs_request_semaphore = BoundedSemaphore(1)

# # Define a function to be run in a thread
# def fetch_interface_data(args):
#     cls, device_name, interface_name = args
#     formatted_inter_name = interface_name.replace('/', '%2F')
#     with ncs_request_semaphore:
#         inter_oper_frag, resp = cls.nso.get_device_live_status(
#             device=device_name,
#             path=f"Cisco-IOS-XR-drivers-media-eth-oper:ethernet-interface/interfaces/interface={formatted_inter_name}"
#         )
#     return inter_oper_frag if inter_oper_frag else []



def split_headers(headers, max_cols):
    split_h = []

    preferred_cols = ["device-name", "interface-name"]
    col_st = len(preferred_cols)  # start splitting after the preferred columns
    cols = max_cols - col_st

    for col_end in range(col_st + cols - 1, len(headers), cols):
        chunk = preferred_cols + headers[col_st:col_end+1]
        split_h.append(chunk)
        col_st = col_end + 1
    return split_h

def generate_markdown_report(headers, devices):
    all_reports = []
    # generating reports
    for hdrs in headers:
        # adding headers to markdown report
        md_report = []
        md_report.append(" | ".join(hdrs))
        md_report.append(" | ".join("---" for _ in hdrs))

        for device in devices:
            device_interfaces = list(Interface.objects.filter(device=device))

            for interface in device_interfaces:
                new_row = [device.name, interface.name, device.device_type.model, device.local_context_data.get("os_version", None), device.site.name]
                md_report.append(" | ".join(new_row))

        all_reports.append(md_report)
    return all_reports


def fetch_device_data(cls, device, split_interface_name, with_nso:bool, timeout:int, retry:int):
    def get_nso_data(path):
        start_time = datetime.now()
        cls.log_info(f"{start_time.strftime('%H:%M:%S')} - Started getting '{path}' for device: '{device.name}' from NSO")
        item_data = {}
        item_data, resp = cls.nso.get_device_live_status(device=device.name, path=path, timeout=timeout, retry=retry)
        end_time = datetime.now()
        time_diff = end_time - start_time
        cls.log_info(f"{end_time.strftime('%H:%M:%S')} - Finished getting '{path}' for device: '{device.name}' from NSO - it took: {time_diff}")
        if not item_data:
            cls.log_warning(f"{path} is empty for device: '{device.name}' url: '{resp.url}'") if cls.with_logs else None
        return item_data, resp
    ############################################################################
    device_interfaces = list(Interface.objects.filter(device=device))
    start_time = datetime.now()
    cls.log_warning(f"{start_time.strftime('%H:%M:%S')} - started reporting for device: '{device.name}'")
    ############################################################################
    device_paths = [
        "ietf-interfaces:interfaces-state",
        "Cisco-IOS-XR-ifmgr-oper:interface-properties/data-nodes",
        "tailf-ned-cisco-ios-xr-stats:controllers/Optics",
        # "Cisco-IOS-XR-drivers-media-eth-oper:ethernet-interface/interfaces"  # disabled as it takes on average 5 mins and is not thread safe...
    ]
    if not with_nso:
        device_paths = []

    state_dict = {}
    property_dict = {}
    optics_dict = {}
    oper_dict = {}

    for path in device_paths:
        try:
            item_data, resp = get_nso_data(path)
        except TimeoutException as e:
            cls.log_failure(f"couldn't retrieve '{path}' due to timeout exception on device: '{device.name}' - {e}")
            continue
        except ConnectionError as e:
            cls.log_failure(f"couldn't retrieve '{path}' due to ConnectionError exception on device: '{device.name}' - {e}")
            continue
        except Exception as e:
            cls.log_failure(f"couldn't retrieve '{path}' due to unhandled exception on device: '{device.name}' - {e}")
            continue
        if path == "ietf-interfaces:interfaces-state":
            if item_data:
                state_dict = {entry['name']: entry for entry in item_data.get('interface', [])}
        elif path == "Cisco-IOS-XR-ifmgr-oper:interface-properties/data-nodes":
            if item_data:
                property_dict = item_data.get('data-node', [{}])[0].get("system-view", {}).get('interfaces', {}).get('interface', {})
                property_dict = {entry['interface-name']: entry for entry in property_dict}

            ######################################################
            # if item_data:
            #     cls.log_warning("item_data not empty")
            # else:
            #     cls.log_warning("item_data empty")
            #     cls.log_warning(f"resp.text:\n{resp.text}")
            # cls.log_failure(resp.status_code)
            # if optics_dict:
            #     cls.log_warning("optics_dict not empty")
            # else:
            #     cls.log_warning("optics_dict empty")
            # raise Exception("end of test")
            ######################################################
        elif path == "tailf-ned-cisco-ios-xr-stats:controllers/Optics":
            if item_data:
                optics_dict = {entry['id']: entry for entry in item_data}
        elif path == "Cisco-IOS-XR-drivers-media-eth-oper:ethernet-interface/interfaces":
            if item_data:
                oper_dict = {entry['interface-name']: entry for entry in item_data.get("interface", [])}
    ############################################################################
    data_rows = []
    # iterate over each interface of the current device
    for interface in device_interfaces:
        current_interface_state = state_dict.get(interface.name, {})
        current_interface_property = property_dict.get(interface.name, {})
        current_interface_oper = oper_dict.get(interface.name, {})

        _, interface_id = split_interface_name(interface.name)
        current_interface_optics = optics_dict.get(interface_id, {})

        # if do_once and interface.name == "Bundle-Ether5":
        #     do_once = False
        #     cls.log_debug(json_dumps_(current_interface_state, indent=4))
        #     cls.log_debug(json_dumps_(current_interface_property, indent=4))
        #     cls.log_debug(json_dumps_(current_interface_oper, indent=4))
        #     cls.log_debug(json_dumps_(current_interface_optics, indent=4))

        if current_interface_state:
            interface.enabled = True if current_interface_state["admin-status"] == "up" else False
            interface.mac_address = current_interface_state.get("phys-address") or None
            interface.speed = int(current_interface_state["speed"]) // 1000000 or None
            try:
                interface.full_clean()
            except ValidationError as e:
                errors = e.message_dict
                if 'speed' in errors:
                    cls.log_failure(f"interface: '{interface.name}' current_interface_state speed: '{current_interface_state['speed']}' must be less or equal to: '2147483647'")
                    interface.speed = None
                else:
                    cls.log_failure(f"Hit unhandled exception: {e} {type(e)}")
                    raise e
            interface.save()

        local_context = device.local_context_data or {}
        try:
            data_rows.append([
                device.name,
                interface.name,
                device.device_type.model,
                local_context.get("os_version", "N/A"),
                device.site.name if device.site else "N/A",
                interface.type,
                interface.description or "N/A",
                "up" if interface.enabled else "down",
                current_interface_state.get("oper-status", "N/A"),
                "up" if current_interface_property.get("line-state") == "im-state-up" else ("down" if current_interface_property.get("line-state") else "N/A"),
                int(current_interface_property.get("bandwidth")) // 1000000 if current_interface_property.get("bandwidth") else "N/A",
                interface.speed or "N/A",
                current_interface_optics.get("instance", {"transceiver-vendor-details": {"optics-type": "N/A"}}).get("transceiver-vendor-details", {"optics-type": "N/A"}).get("optics-type", "N/A"),
                current_interface_optics.get("instance", {"transceiver-vendor-details": {"part-number": "N/A"}}).get("transceiver-vendor-details", {"part-number": "N/A"}).get("part-number", "N/A"),
                current_interface_oper.get("phy-details", {"optics-wavelength": "N/A"}).get("optics-wavelength", "N/A"),
                interface.mtu or "N/A",
                current_interface_property.get("ietf-ip:ipv4", {"mtu": "N/A"}).get("mtu", "N/A"),
                ", ".join(str(ip) for ip in interface.ip_addresses.all() if ip.family == 4) if hasattr(interface, 'ip_addresses') and any(ip.family == 4 for ip in interface.ip_addresses.all()) else "N/A",
                ", ".join(str(ip) for ip in interface.ip_addresses.all() if ip.family == 6) if hasattr(interface, 'ip_addresses') and any(ip.family == 6 for ip in interface.ip_addresses.all()) else "N/A",
                str(interface.mac_address) or "N/A",
                len(interface.member_interfaces.all()) if hasattr(interface, 'member_interfaces') and interface.type == "lag" else "N/A",
                interface.lag.name if interface.lag and interface.type != "lag" else "N/A",
                interface.mode or "N/A",
                interface.untagged_vlan or "N/A",
                ", ".join(str(vlan.id) for vlan in list(interface.tagged_vlans.all())) or "N/A",
                interface.vrf.name if interface.vrf else "default",
                interface.connected_endpoints[0].device.name if interface.connected_endpoints else "N/A",
                interface.connected_endpoints[0].name if interface.connected_endpoints else "N/A",
                interface.connected_endpoints[0].lag.name if interface.connected_endpoints and interface.connected_endpoints[0].lag else "N/A",
                (local_context.get("interfaces", {}).get(interface.name, {}).get("service-policy", {}).get("input-list", []) or [{"name": "N/A"}])[0].get("name"),
                (local_context.get("interfaces", {}).get(interface.name, {}).get("service-policy", {}).get("output-list", []) + [{"name": "N/A"}, {"name": "N/A"}])[0:2][0].get('name'),
                (local_context.get("interfaces", {}).get(interface.name, {}).get("service-policy", {}).get("output-list", []) + [{"name": "N/A"}, {"name": "N/A"}])[0:2][1].get('name'),
            ])
        except Exception as e:
            end_time = datetime.now()
            time_diff = end_time - start_time
            cls.log_failure(f"{end_time.strftime('%H:%M:%S')} - interface.name: '{interface.name}' device: '{device.name}' - {e} ")
            raise
    end_time = datetime.now()
    time_diff = end_time - start_time
    cls.log_warning(f"{end_time.strftime('%H:%M:%S')} - Finished reporting for device: '{device.name}'")
    return data_rows


def generate_excel_report(cls, wb, headers, devices, reports_dir, split_interface_name, timeout:int, retry:int, with_nso:bool, report_name="report.xlsx"):
    # creating a new workbook and selecting the active sheet
    ws = wb.active

    # write the headers in the first row
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)

    # initialize row index for data
    row_idx = 2

    # Use a ThreadPoolExecutor to parallelize the operation
    with ThreadPoolExecutor(max_workers=5) as executor:
        # Map devices to the fetch_device_data function
        results = list(executor.map(lambda device: fetch_device_data(cls, device, split_interface_name, with_nso=with_nso, timeout=timeout, retry=retry), devices))

    # After all threads have completed, results is a list of all data_rows
    for data_rows in results:
        for data_row in data_rows:
            for col_idx, item in enumerate(data_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=item)
            row_idx += 1
    # save the workbook to a file
    wb.save(f"{reports_dir}/{report_name}")



class GenerateReport(Script):
    class Meta:
        name = "Generates the CSG services inventory"
        description = """
            Generates a CSV inventory file from the onboarded CSG devices
        """
        job_timeout = 432000  # 5 days
        commit_default = True
        commit = False
        scheduling_enabled = False

    limit = IntegerVar(
        required=True,
        default=1
    )

    offset = IntegerVar(
        required=True,
        default="0"
    )

    devices = TextVar(
        required=False,
    )

    with_logs = BooleanVar(
        default=True,
    )

    with_nso = BooleanVar(
        default=True,
    )

    base_url = StringVar(
        required=True,
        default="10.0.27.6:8080",
    )

    username = StringVar(
        required=True,
        default="ifoughal",
    )

    password = StringVar(
        required=True,
        widget=PasswordInput,
    )

    nso_timeout = IntegerVar(
        required=True,
        default=500
    )

    nso_retry = IntegerVar(
        required=True,
        default=1
    )

    with_multithreading = BooleanVar(
        default=True,
    )


    def run(self, data, commit):
        try:
            start_time = datetime.now()
            self.log_info(f"{start_time.strftime('%H:%M:%S')} - Started reporting script")
            headers = [
                "device-name", "interface-name", "device-model", "device-version", "device-location",
                "interface-type", "interface-description", "interface-admin-state", "interface-oper-state",
                "interface-link-state", "bandwidth", "speed (mbps)", "optics-type", "optics-part_number",
                "optics-WaveLength", "MTU", "MTU-IP", "ipv4-address", "ipv6-address", "interface-mac_address",
                "members-count", "parent-interface", "802.1Q-mode", "untagged-vlan", "tagged-vlan(s)", "VRF", "peer-name",
                "peer-interface", "peer-interface-parent", "Service-Policy Input",
                "Service-Policy Output_1", "Service-Policy Output_2",
            ]
            ##########################################################################################
            from common.utils.device import split_interface_name
            from common.utils.device import DeviceManager
            from common.utils.nso import Nso
            from openpyxl import Workbook
            from common.utils.functions import ThreadPoolExecutorStackTraced
            ##########################################################################################
            with_nso = data.get("with_nso")


            ##########################################################################################
            # instantiate DeviceManager for data parsing and onboarding
            dm = DeviceManager(
                None,
                data["with_logs"],
                [
                    self.log_info,
                    self.log_warning,
                    self.log_failure,
                    self.log_debug
                ]
            )
            ###########################################################################################
            self.nso = None
            if with_nso:
                self.nso = Nso(
                    base_url=data.get('base_url'),
                    username=data.get('username'),
                    password=data.get('password'),
                    log=[
                        self.log_info,
                        self.log_warning,
                        self.log_failure,
                        self.log_debug
                    ],
                )
                result = self.nso.test_credentials()
                if not result:
                    raise AbortScript("failed to authenticate to NSO with current credentials...")
                self.log_success("NSO given crendentials successfuly authenticated!")
            ###########################################################################################
            # Getting devices from netbox
            self.log_info(f"{datetime.now().strftime('%H:%M:%S')} - Getting devices from Netbox")
            try:
                limit_devices = []
                if data.get('devices'):
                    limit_devices = data.get('devices').split(" ")
                nb_devices = dm.get_or_create_csg_devices(limit_devices=limit_devices, limit=data.get('limit'), offset=data.get('offset'))
            except NSODevicesRetrievalError as e:
                raise AbortScript(f"{e}")
            if not nb_devices:
                raise AbortScript(f"failed to retrieve devices from netbox with entered parameteres: limit_devices='{limit_devices}' - limit={data.get('limit')} - offset={data.get('offset')}")
            self.log_info(f"{datetime.now().strftime('%H:%M:%S')} - Retrieved: '{len(nb_devices)}' devices from Netbox.")
            ###########################################################################################
            # TODO: change generated-configs volume to generated/configs
            reports_dir = f"{getcwd()}/generated-configs/reports"
            if not os_path.exists(reports_dir):
                makedirs(reports_dir)

            self.log_info(f"Excel reports will be dumped at: '{reports_dir}'")
            wb = Workbook()
            generate_excel_report(
                wb=wb,
                cls=self,
                headers=headers,
                devices=nb_devices,
                reports_dir=reports_dir,
                report_name="generated_report.xlsx",
                timeout=data.get("nso_timeout"),
                retry=data.get("nso_retry"),
                with_nso=with_nso,
                split_interface_name=split_interface_name
            )
            # headers = split_headers(headers, 5)
            # reports = generate_markdown_report(
            #     headers=headers,
            #     devices=nb_devices
            # )

            # self.log_info(f"Generated markdown report:")
            # for report in reports:
            #     self.log_info('\n'.join(report))

            end_time = datetime.now()
            time_diff = end_time - start_time
            self.log_info(f"{end_time.strftime('%H:%M:%S')} - Finished reporting script - it took: {time_diff}")

        except AbortScript as e:
            raise AbortScript(f"{e}")
        except ModuleNotFoundError as e:
            raise AbortScript(f"Missing module(s): {e}")
        except Exception as e_1:
            error_msg = str(exc_info()[0](format_exc())).split(',')
            error_msg = "```\n" + ''.join(error_msg) + "\n```"
            self.log_failure(error_msg)
            raise AbortScript(f"failed due to caughting unhandled exception")
